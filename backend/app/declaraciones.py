"""
Declaraciones y Pagos — parseo de reportes SUNAT y match declarado vs pagado.

Reportes soportados:
- DETALLE PDT 621 (xlsx): 1 fila por mes; desde la columna G pares (nro casilla, valor).
  Casillas clave: 188 = IGV total deuda, 324 = Renta total deuda.
- DETALLE DE DECLARACIONES Y PAGOS (xls legado o xlsx): 1 fila por pago con
  periodo YYYYMM, formulario, banco, código de tributo e importe pagado.
"""
import io
import json

MESES = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio",
         "Agosto", "Setiembre", "Octubre", "Noviembre", "Diciembre"]
MES_NUM = {m.lower(): i + 1 for i, m in enumerate(MESES)}
MES_NUM["septiembre"] = 9

# Casillas del PDT 621 expuestas como columnas editables
CASILLAS_CLAVE = {
    100: "ventas_base", 101: "ventas_igv",
    107: "compras_base", 108: "compras_igv",
    301: "renta_ingresos", 312: "renta_pago_cta",
    140: "igv_resultante", 184: "igv_a_pagar", 304: "renta_neta",
    188: "igv_deuda", 324: "renta_deuda",
}

# Código de tributo SUNAT → categoría
CATEGORIA_TRIBUTO = {
    "1011": "IGV", "1012": "IGV", "1016": "IGV",
    "3031": "RENTA3", "3032": "RENTA3", "3033": "RENTA3",
    "3041": "RENTA4", "3042": "RENTA4",
    "3051": "RENTA5", "3052": "RENTA5",
    "3081": "RENTA_ANUAL",
    "5310": "ONP", "5340": "ONP",
    "5210": "ESSALUD", "5211": "ESSALUD", "5222": "ESSALUD",
    "8021": "FRACC", "8026": "FRACC", "8027": "FRACC", "8028": "FRACC",
}
CATEGORIA_LABEL = {
    "IGV": "IGV", "RENTA3": "Renta 3ra", "RENTA4": "Renta 4ta",
    "RENTA5": "Renta 5ta", "RENTA_ANUAL": "Renta Anual", "ONP": "ONP",
    "ESSALUD": "EsSalud", "FRACC": "Fraccionamiento", "OTROS": "Otros",
}
# Las categorías que amortizan la deuda del PDT 621 (match declarado vs pagado)
CATEGORIAS_PDT = ("IGV", "RENTA3")


def _f(v):
    """Número o None (tolera strings con coma de miles)."""
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


def _s(v):
    return str(v).strip() if v is not None else ""


def parsear_pdt621(file_bytes: bytes) -> list:
    """Devuelve 1 dict por mes declarado (meses sin datos se omiten)."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    out = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            if not row or len(row) < 8:
                continue
            anio, mes = row[0], MES_NUM.get(_s(row[1]).lower())
            if not isinstance(anio, (int, float)) or not mes:
                continue
            casillas = {}
            for i in range(6, len(row) - 1, 2):
                cas, val = row[i], row[i + 1]
                if isinstance(cas, (int, float)) and 100 <= int(cas) <= 999:
                    casillas[int(cas)] = _f(val)
            if not any(v is not None for v in casillas.values()):
                continue  # mes aún no declarado
            item = {"anio": int(anio), "mes": mes,
                    "tipo_decl": _s(row[3]) or "Original",
                    "igv_justo": _s(row[4]),
                    "detalle_json": json.dumps(casillas)}
            for cas, campo in CASILLAS_CLAVE.items():
                item[campo] = casillas.get(cas)
            out.append(item)
        if out:
            break
    return out


def _leer_filas(file_bytes: bytes) -> list:
    """Lee xls legado (xlrd) o xlsx (openpyxl) → lista de listas."""
    if file_bytes[:4] == b"\xd0\xcf\x11\xe0":  # OLE2 = .xls legado
        import xlrd
        wb = xlrd.open_workbook(file_contents=file_bytes)
        ws = wb.sheet_by_index(0)
        filas = []
        for r in range(ws.nrows):
            fila = []
            for c in range(ws.ncols):
                cell = ws.cell(r, c)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    dt = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                    fila.append(dt.strftime("%d/%m/%Y"))
                else:
                    fila.append(cell.value)
            filas.append(fila)
        return filas
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    return [list(r) for r in wb.active.iter_rows(values_only=True)]


def _fecha_iso(v) -> str:
    s = _s(v)
    if "/" in s:  # dd/mm/yyyy
        p = s.split("/")
        if len(p) == 3:
            return f"{p[2]}-{p[1].zfill(2)}-{p[0].zfill(2)}"
    return s[:10]


def _cod_tributo(v) -> str:
    s = _s(v)
    if s in ("-", ""):
        return ""
    try:
        return str(int(float(s)))
    except ValueError:
        return s


def parsear_pagos(file_bytes: bytes) -> list:
    """Devuelve 1 dict por pago del reporte 'Detalle de declaraciones y pagos'."""
    filas = _leer_filas(file_bytes)
    header, col = None, {}
    for idx, r in enumerate(filas[:40]):
        low = [_s(v).lower() for v in r]
        if "periodo" in low and any("importe" in x for x in low):
            desc_cols = []
            for ci, x in enumerate(low):
                if x == "periodo":
                    col["periodo"] = ci
                elif "formulario" in x:
                    col["formulario"] = ci
                elif "orden" in x:
                    col["orden"] = ci
                elif "descripcion" in x or "descripción" in x:
                    desc_cols.append(ci)
                elif "banco" in x:
                    col["banco"] = ci
                elif "fecha" in x:
                    col["fecha"] = ci
                elif "tributo" in x:
                    col["cod_tributo"] = ci
                elif "importe" in x:
                    col["importe"] = ci
            if desc_cols:
                col["descripcion"] = desc_cols[0]
                if len(desc_cols) > 1:
                    col["tributo"] = desc_cols[-1]
            header = idx
            break
    if header is None:
        return []

    def _get(r, key):
        ci = col.get(key)
        return r[ci] if ci is not None and ci < len(r) else None

    out = []
    for r in filas[header + 1:]:
        periodo = _s(_get(r, "periodo")).replace(".0", "")
        if len(periodo) != 6 or not periodo.isdigit():
            continue
        cod = _cod_tributo(_get(r, "cod_tributo"))
        out.append({
            "anio": int(periodo[:4]), "mes": int(periodo[4:6]),
            "formulario": _s(_get(r, "formulario")).replace(".0", ""),
            "orden": _s(_get(r, "orden")).replace(".0", ""),
            "descripcion": _s(_get(r, "descripcion"))[:120],
            "banco": _s(_get(r, "banco"))[:40],
            "fecha_pago": _fecha_iso(_get(r, "fecha")),
            "cod_tributo": cod,
            "tributo": _s(_get(r, "tributo"))[:120],
            "categoria": CATEGORIA_TRIBUTO.get(cod, "OTROS"),
            "importe": _f(_get(r, "importe")) or 0,
        })
    return out


def construir_reporte(decls: list, pagos: list, anio: int) -> dict:
    """Match declarado (PDT 621) vs pagado por mes + matriz por tributo."""
    por_mes_d = {d.mes: d for d in decls}
    meses = []
    tot_declarado = tot_pagado_pdt = tot_pagado_all = tot_pendiente = 0.0
    categorias_presentes = set()
    pendientes = []
    for mes in range(1, 13):
        d = por_mes_d.get(mes)
        pagos_mes = [p for p in pagos if p.mes == mes]
        matriz = {}
        for p in pagos_mes:
            matriz[p.categoria] = matriz.get(p.categoria, 0) + (p.importe or 0)
            if (p.importe or 0) > 0:
                categorias_presentes.add(p.categoria)
        declarado_igv = (d.igv_deuda or 0) if d else 0
        declarado_renta = (d.renta_deuda or 0) if d else 0
        declarado = declarado_igv + declarado_renta
        pagado_pdt = sum(matriz.get(c, 0) for c in CATEGORIAS_PDT)
        pagado_all = sum(matriz.values())
        diferencia = declarado - pagado_pdt
        if d is None:
            estado = "sin_declarar"
        elif declarado <= 0 or diferencia <= 0:
            estado = "pagado" if pagado_pdt > 0 else "al_dia"
        elif pagado_pdt > 0:
            estado = "parcial"
        else:
            estado = "pendiente"
        if estado in ("parcial", "pendiente"):
            pendientes.append({"mes": mes, "nombre": MESES[mes - 1],
                               "monto": round(diferencia, 2)})
            tot_pendiente += diferencia
        tot_declarado += declarado
        tot_pagado_pdt += pagado_pdt
        tot_pagado_all += pagado_all
        meses.append({
            "mes": mes, "nombre": MESES[mes - 1],
            "declarado_igv": declarado_igv, "declarado_renta": declarado_renta,
            "declarado": declarado, "pagado_pdt": pagado_pdt,
            "pagado_total": pagado_all,
            "diferencia": round(diferencia, 2) if d else 0,
            "estado": estado, "matriz": matriz,
            "pagos": [{"id": p.id, "fecha": p.fecha_pago, "formulario": p.formulario,
                       "descripcion": p.descripcion, "banco": p.banco,
                       "tributo": p.tributo, "categoria": p.categoria,
                       "importe": p.importe} for p in pagos_mes],
        })
    return {
        "anio": anio, "meses": meses,
        "totales": {"declarado": round(tot_declarado, 2),
                    "pagado_pdt": round(tot_pagado_pdt, 2),
                    "pagado_total": round(tot_pagado_all, 2),
                    "pendiente": round(max(tot_pendiente, 0), 2)},
        "pendientes": pendientes,
        "categorias": [c for c in CATEGORIA_LABEL if c in categorias_presentes],
        "categoria_labels": CATEGORIA_LABEL,
    }
