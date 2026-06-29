"""
Lógica de Planilla (PLAME): trabajadores (5ta) y renta de 4ta.
Incluye serialización, cálculo de totales e importación/exportación Excel.
"""
import io


def _num(v) -> float:
    try:
        return round(float(v or 0), 2)
    except Exception:
        return 0.0


def trab_dict(t, empresa_nombre: str = "") -> dict:
    ingresos = _num(t.remuneracion) + _num(t.asignacion_familiar) + _num(t.otros_ingresos)
    descuentos = _num(t.aporte_pension) + _num(t.renta_quinta) + _num(t.otros_descuentos)
    return {
        "id": t.id, "empresa_id": t.empresa_id, "empresa": empresa_nombre,
        "periodo": t.periodo or "", "num_doc": t.num_doc or "", "nombre": t.nombre or "",
        "regimen_pensionario": t.regimen_pensionario or "ONP", "afp_nombre": t.afp_nombre or "",
        "dias_laborados": t.dias_laborados or 0,
        "remuneracion": _num(t.remuneracion), "asignacion_familiar": _num(t.asignacion_familiar),
        "otros_ingresos": _num(t.otros_ingresos), "aporte_pension": _num(t.aporte_pension),
        "essalud": _num(t.essalud), "renta_quinta": _num(t.renta_quinta),
        "otros_descuentos": _num(t.otros_descuentos),
        "total_ingresos": round(ingresos, 2), "total_descuentos": round(descuentos, 2),
        "neto_pagar": round(ingresos - descuentos, 2),
    }


def cuarta_dict(c, empresa_nombre: str = "") -> dict:
    return {
        "id": c.id, "empresa_id": c.empresa_id, "empresa": empresa_nombre,
        "periodo": c.periodo or "", "tipo_doc": c.tipo_doc or "RUC", "num_doc": c.num_doc or "",
        "nombre": c.nombre or "", "num_recibo": c.num_recibo or "", "fecha_emision": c.fecha_emision or "",
        "monto_bruto": _num(c.monto_bruto), "retencion": _num(c.retencion),
        "neto_pagar": round(_num(c.monto_bruto) - _num(c.retencion), 2),
    }


# ── Exportación Excel ─────────────────────────────────────────────────────────
TRAB_COLS = [
    ("num_doc", "N° Documento"), ("nombre", "Apellidos y Nombres"),
    ("regimen_pensionario", "Régimen pensión"), ("afp_nombre", "AFP"),
    ("dias_laborados", "Días lab."), ("remuneracion", "Remuneración"),
    ("asignacion_familiar", "Asig. familiar"), ("otros_ingresos", "Otros ingresos"),
    ("total_ingresos", "Total ingresos"), ("aporte_pension", "Aporte pensión"),
    ("essalud", "EsSalud 9%"), ("renta_quinta", "Renta 5ta"),
    ("otros_descuentos", "Otros desc."), ("neto_pagar", "Neto a pagar"),
]
CUARTA_COLS = [
    ("tipo_doc", "Tipo doc."), ("num_doc", "N° Documento"), ("nombre", "Razón social / Nombre"),
    ("num_recibo", "N° Recibo (RHE)"), ("fecha_emision", "Fecha emisión"),
    ("monto_bruto", "Monto bruto"), ("retencion", "Retención 8%"), ("neto_pagar", "Neto"),
]


COLAB_COLS = [
    ("tipo_doc", "Tipo doc."), ("num_doc", "N° Documento"), ("ap_paterno", "Ap. paterno"),
    ("ap_materno", "Ap. materno"), ("nombres", "Nombres"), ("fecha_nacimiento", "F. nacimiento"),
    ("sexo", "Sexo"), ("nacionalidad", "Nacionalidad"), ("tipo_trabajador", "Tipo trabajador"),
    ("regimen_laboral", "Régimen laboral"), ("tipo_contrato", "Tipo contrato"),
    ("ocupacion", "Ocupación"), ("jornada", "Jornada"), ("fecha_ingreso", "F. ingreso"),
    ("situacion", "Situación"), ("regimen_pensionario", "Pensión"), ("afp_nombre", "AFP"),
    ("cuspp", "CUSPP"), ("regimen_salud", "Salud"), ("remuneracion", "Remuneración"),
    ("tipo_pago", "Forma pago"), ("cci", "CCI"), ("email", "Correo"), ("telefono", "Teléfono"),
]


def exportar_excel(filas: list, cols: list, titulo: str, periodo: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "PLAME"
    ws["A1"] = f"{titulo} — Período {periodo or 'todos'}"
    ws["A1"].font = Font(bold=True, size=13, color="1B3A6B")
    ws.append([])
    header = [c[1] for c in cols]
    ws.append(header)
    hfill = PatternFill("solid", fgColor="1B3A6B")
    for cell in ws[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center")
    for fila in filas:
        ws.append([fila.get(k, "") for k, _ in cols])
    for i, (k, _) in enumerate(cols, 1):
        ws.column_dimensions[ws.cell(row=3, column=i).column_letter].width = 16
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Importación Excel ─────────────────────────────────────────────────────────
def _norm(s: str) -> str:
    return (str(s or "")).strip().lower().replace("°", "").replace(".", "").replace("  ", " ")


def importar_excel(file_bytes: bytes, cols: list) -> list:
    """Lee un xlsx y mapea filas por encabezado (tolerante). Devuelve lista de dicts."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    # Buscar la fila de encabezados (la que más coincida con los nombres esperados)
    etiquetas = {_norm(lbl): key for key, lbl in cols}
    etiquetas.update({_norm(key): key for key, _ in cols})
    header_idx, mapeo = None, {}
    for idx, r in enumerate(rows[:6]):
        m = {}
        for ci, val in enumerate(r):
            k = etiquetas.get(_norm(val))
            if k:
                m[ci] = k
        if len(m) >= 2:
            header_idx, mapeo = idx, m
            break
    if header_idx is None:
        return []
    out = []
    for r in rows[header_idx + 1:]:
        item = {}
        for ci, key in mapeo.items():
            if ci < len(r):
                item[key] = r[ci]
        if any(str(v).strip() for v in item.values() if v is not None):
            out.append(item)
    return out
